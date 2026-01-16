#!/usr/bin/env python3
import os
import threading
import warnings
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import pandas as pd
import pyodbc
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.comments import Comment

warnings.filterwarnings("ignore", category=UserWarning)

# -------------------------------
# Appearance detection (macOS)
# -------------------------------
def is_dark_mode():
    try:
        import subprocess
        r = subprocess.run(
            ["defaults", "read", "-g", "AppleInterfaceStyle"],
            capture_output=True,
            text=True
        )
        return "Dark" in r.stdout
    except Exception:
        return False


# -------------------------------
# Themes
# -------------------------------
THEMES = {
    "dark": {
        "bg": "#1e1e1e",
        "fg": "#e6e6e6",
        "entry": "#2b2b2b",
        "accent": "#4a90e2",
    },
    "light": {
        "bg": "#f4f4f4",
        "fg": "#1e1e1e",
        "entry": "#ffffff",
        "accent": "#007aff",
    }
}

# -------------------------------
# Field Mapping Links
# -------------------------------
MAPPING_LINKS = {
    "1-Loans": "https://help.themortgageoffice.com/knowledge/loan-field-mappings",
    "2-Co-Borrowers": "https://help.themortgageoffice.com/knowledge/co-borrower-field-mappings",
    "3-Fundings": "https://help.themortgageoffice.com/knowledge/what-are-the-vendor-field-mappings",
    "4-Properties_&_Insurance": "https://help.themortgageoffice.com/knowledge/what-are-the-properties-insurance-field-mappings",
}

# -------------------------------
# Helpers
# -------------------------------
def get_field_definitions(url):
    defs = {}
    if not url:
        return defs
    try:
        soup = BeautifulSoup(requests.get(url, timeout=10).text, "html.parser")
        for row in soup.find_all("tr"):
            cells = row.find_all(["td", "th"])
            if len(cells) >= 2:
                key = cells[0].get_text(strip=True)
                desc = cells[1]
                for tag in desc.find_all(["br", "p", "li"]):
                    tag.insert_before("\n")
                defs[key] = "\n".join(
                    l.strip() for l in desc.get_text().splitlines() if l.strip()
                )
    except Exception:
        pass
    return defs


def clean_and_format_df(df, omissions):
    df = df.drop(columns=[c for c in df.columns if "recid" in c.lower()], errors="ignore")
    df = df.drop(columns=[c for c in omissions if c in df.columns], errors="ignore")

    invalid = [0, "0", "0.0", "", " ", None, "NULL", "null"]
    df = df.loc[:, ~(df.isin(invalid)).all(axis=0)]

    for col in df.columns:
        if "date" in col.lower():
            try:
                df[col] = (
                    pd.to_datetime(df[col], errors="coerce")
                    .dt.strftime("%m/%d/%Y")
                    .fillna("")
                )
            except Exception:
                pass
    return df


# -------------------------------
# SQL connections
# -------------------------------
def connect(server, database):
    return pyodbc.connect(
        f"Driver={{ODBC Driver 18 for SQL Server}};"
        f"Server={server};Database={database};"
        "Authentication=ActiveDirectoryInteractive;"
        "Encrypt=yes;TrustServerCertificate=yes;"
    )


def search_databases(term):
    conn = connect("10.1.7.5", "AbsWebSys")
    query = """
        SELECT
            RecID,
            Description,
            DatabaseName,
            Server
        FROM AbsWebSys.dbo.CompanyDatabase
        WHERE DatabaseName LIKE ?
        ORDER BY Description
    """
    df = pd.read_sql(query, conn, params=[f"%{term}%"])
    conn.close()
    return df


# -------------------------------
# Export logic
# -------------------------------
def run_export(server, database, folder, log, set_status):
    try:
        set_status("Connecting to target database…")
        conn = connect(server, database)

        os.makedirs(folder, exist_ok=True)
        omissions = [
            "SysCreatedDate", "SysTimeStamp", "SysRecStatus", "SysCreatedBy",
            "BorrowerF", "CompanyF", "SortName", "ByLastName", "FundControl",
            "ACH_IndividualId", "XML", "LanguagePreference", "InsuranceDocument"
        ]

        def export(name, query):
            set_status(f"Exporting {name}…")
            df = clean_and_format_df(pd.read_sql(query, conn), omissions)
            path = os.path.join(folder, f"{name}.xlsx")
            df.to_excel(path, index=False)

            wb = load_workbook(path)
            ws = wb.active
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = max(
                    len(str(c.value)) if c.value else 0 for c in col
                ) + 3

            defs = get_field_definitions(MAPPING_LINKS.get(name))
            for cell in ws[1]:
                key = str(cell.value)
                if key in defs:
                    cell.comment = Comment(defs[key], "TMO")
            wb.save(path)
            log(f"✓ {name}")

        export("1-Loans", "SELECT * FROM [TDS Loans]")
        export(
            "2-Co-Borrowers",
            "SELECT l.Account, t.* FROM [TDS CoBorrowers] t "
            "LEFT JOIN [TDS Loans] l ON t.LoanRecID = l.RecID"
        )
        export(
            "3-Fundings",
            """
            SELECT l.Account AS Loan_Account, lend.Account AS Lender_Account, f.*, d.*
            FROM [TDS Funding] f
            INNER JOIN [TDS Draws] d ON f.RecID = d.FundingRecID
            LEFT JOIN [TDS Loans] l ON f.LoanRecID = l.RecID
            LEFT JOIN [TDS Lenders] lend ON f.LenderRecID = lend.RecID
            """
        )

        p_raw = pd.read_sql(
            """
            SELECT l.Account, p.RecID AS _pid, p.*
            FROM [TDS Properties] p
            LEFT JOIN [TDS Loans] l ON p.LoanRecID = l.RecID
            """,
            conn
        )

        p_clean = clean_and_format_df(p_raw.drop(columns=["Account", "_pid"]), omissions)
        p_clean.columns = [f"Property: {c}" for c in p_clean.columns]
        p_clean["Account"] = p_raw["Account"]
        p_clean["_pid"] = p_raw["_pid"]

        i_raw = pd.read_sql("SELECT * FROM [TDS Insurance]", conn)
        i_clean = clean_and_format_df(i_raw, omissions)
        i_clean.columns = [f"Insurance: {c}" for c in i_clean.columns]
        i_clean["_pref"] = i_raw["PropRecID"]

        df4 = (
            pd.merge(p_clean, i_clean, left_on="_pid", right_on="_pref", how="left")
            .drop(columns=["_pid", "_pref"])
        )

        path = os.path.join(folder, "4-Properties_&_Insurance.xlsx")
        df4.to_excel(path, index=False)

        conn.close()
        set_status("Completed")
        messagebox.showinfo("Success", "Production export completed.")

    except Exception as e:
        set_status("Error")
        messagebox.showerror("Error", str(e))


# -------------------------------
# UI
# -------------------------------
def start_app():
    root = tk.Tk()
    root.title("TDS Excel Production Suite")
    root.geometry("820x620")

    theme = "dark" if is_dark_mode() else "light"
    colors = THEMES[theme]

    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure(".", background=colors["bg"], foreground=colors["fg"])
    style.configure("TEntry", fieldbackground=colors["entry"])
    style.configure("Primary.TButton", background=colors["accent"], foreground="white")

    root.configure(bg=colors["bg"])
    frame = ttk.Frame(root, padding=15)
    frame.pack(fill="both", expand=True)

    # Search
    ttk.Label(frame, text="Search Database").pack(anchor="w")
    search = ttk.Entry(frame, width=50)
    search.pack(anchor="w", pady=4)

    results = ttk.Treeview(
        frame,
        columns=("Description", "Database", "Server"),
        show="headings",
        height=8
    )
    for col in results["columns"]:
        results.heading(col, text=col)
        results.column(col, width=240 if col != "Server" else 120)

    results.pack(fill="x", pady=8)

    server_var = tk.StringVar()
    db_var = tk.StringVar()

    def do_search():
        results.delete(*results.get_children())
        df = search_databases(search.get())
        for _, r in df.iterrows():
            results.insert("", "end", values=(r.Description, r.DatabaseName, r.Server))

    ttk.Button(frame, text="Search", command=do_search).pack(anchor="w")

    def on_select(_):
        item = results.selection()
        if item:
            vals = results.item(item[0], "values")
            db_var.set(vals[1])
            server_var.set(vals[2])

    results.bind("<<TreeviewSelect>>", on_select)

    # Selected info
    ttk.Label(frame, text="Selected Server").pack(anchor="w", pady=(10, 0))
    ttk.Entry(frame, textvariable=server_var, state="readonly").pack(fill="x")

    ttk.Label(frame, text="Selected Database").pack(anchor="w", pady=(10, 0))
    ttk.Entry(frame, textvariable=db_var, state="readonly").pack(fill="x")

    ttk.Label(frame, text="Output Folder").pack(anchor="w", pady=(10, 0))
    folder = ttk.Entry(frame)
    folder.insert(0, os.getcwd())
    folder.pack(fill="x")

    ttk.Button(
        frame,
        text="Browse",
        command=lambda: folder.delete(0, tk.END)
        or folder.insert(0, filedialog.askdirectory())
    ).pack(anchor="w", pady=4)

    status = tk.StringVar(value="Ready")
    ttk.Label(frame, textvariable=status).pack(anchor="w", pady=(10, 2))

    progress = ttk.Progressbar(frame, mode="indeterminate")
    progress.pack(fill="x")

    logbox = tk.Text(frame, height=8, bg=colors["entry"], fg=colors["fg"])
    logbox.pack(fill="both", pady=6)

    def log(msg):
        logbox.insert("end", msg + "\n")
        logbox.see("end")

    def set_status(msg):
        status.set(msg)

    def run():
        progress.start()
        threading.Thread(
            target=lambda: (
                run_export(server_var.get(), db_var.get(), folder.get(), log, set_status),
                progress.stop()
            ),
            daemon=True
        ).start()

    ttk.Button(
        frame,
        text="RUN PRODUCTION EXPORT",
        style="Primary.TButton",
        command=run
    ).pack(pady=10)

    root.mainloop()


if __name__ == "__main__":
    start_app()