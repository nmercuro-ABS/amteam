#!/usr/bin/env python3
import pandas as pd
import pyodbc
import os
import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl.comments import Comment
from openpyxl import load_workbook
import warnings

warnings.filterwarnings('ignore', category=UserWarning)

# -------------------------------
# Field Mapping Links
# -------------------------------
MAPPING_LINKS = {
    '1-Loans': "https://help.themortgageoffice.com/knowledge/loan-field-mappings",
    '2-Co-Borrowers': "https://help.themortgageoffice.com/knowledge/co-borrower-field-mappings",
    '3-Fundings': "https://help.themortgageoffice.com/knowledge/what-are-the-vendor-field-mappings",
    '4-Properties_&_Insurance': "https://help.themortgageoffice.com/knowledge/what-are-the-properties-insurance-field-mappings"
}


# -------------------------------
# Helper Functions (UNCHANGED)
# -------------------------------
def get_field_definitions(url):
    defs = {}
    if not url:
        return defs
    try:
        r = requests.get(url, timeout=10)
        soup = BeautifulSoup(r.text, "html.parser")
        for row in soup.find_all("tr"):
            cells = row.find_all(["td", "th"])
            if len(cells) >= 2:
                field = cells[0].get_text(strip=True)
                desc = cells[1]
                for tag in desc.find_all(["br", "p", "li"]):
                    tag.insert_before("\n")
                lines = [l.strip() for l in desc.get_text().splitlines() if l.strip()]
                defs[field] = "\n".join(lines)
    except:
        pass
    return defs


def clean_and_format_df(df, omissions):
    recid_cols = [c for c in df.columns if 'recid' in c.lower()]
    df = df.drop(columns=recid_cols, errors="ignore")
    df = df.drop(columns=[c for c in df.columns if c in omissions], errors="ignore")

    invalid = [0, 0.0, '0', '0.0', '', ' ', 'NULL', 'null', None]
    df = df.loc[:, ~(df.isin(invalid)).all(axis=0)]

    for col in df.columns:
        if 'date' in col.lower():
            try:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%#m/%#d/%Y')
                df[col] = df[col].replace('NaT', '')
            except:
                pass
    return df


# -------------------------------
# Main Export Logic (UNCHANGED)
# -------------------------------
def run_production_export(server, database, output_folder):
    try:
        OMISSIONS = [
            'SysCreatedDate', 'SysTimeStamp', 'SysRecStatus', 'SysCreatedBy',
            'BorrowerF', 'CompanyF', 'SortName', 'ByLastName', 'FundControl',
            'ACH_IndividualId', 'XML', 'LanguagePreference', 'InsuranceDocument'
        ]

        conn_str = (
            f"Driver={{ODBC Driver 17 for SQL Server}};"
            f"Server={server};"
            f"Database={database};"
            "Authentication=ActiveDirectoryInteractive;"
            "Encrypt=yes;TrustServerCertificate=yes;"
        )
        conn = pyodbc.connect(conn_str)

        os.makedirs(output_folder, exist_ok=True)

        df1 = pd.read_sql("SELECT * FROM [TDS Loans]", conn)
        df1 = clean_and_format_df(df1, OMISSIONS)
        df1["ReserveBalance"] = ""
        df1["ImpoundBalance"] = ""

        df2 = pd.read_sql(
            "SELECT l.Account, t.* FROM [TDS CoBorrowers] t "
            "LEFT JOIN [TDS Loans] l ON t.LoanRecID = l.RecID", conn
        )
        df2 = clean_and_format_df(df2, OMISSIONS)

        f_query = """
            SELECT l.Account AS Loan_Account, lend.Account AS Lender_Account, f.*, d.*
            FROM [TDS Funding] f
            INNER JOIN [TDS Draws] d ON f.RecID = d.FundingRecID
            LEFT JOIN [TDS Loans] l ON f.LoanRecID = l.RecID
            LEFT JOIN [TDS Lenders] lend ON f.LenderRecID = lend.RecID
        """
        df3 = clean_and_format_df(pd.read_sql(f_query, conn), OMISSIONS)

        p_raw = pd.read_sql(
            "SELECT l.Account, p.RecID as _pid, p.* "
            "FROM [TDS Properties] p "
            "LEFT JOIN [TDS Loans] l ON p.LoanRecID = l.RecID", conn
        )
        p_clean = clean_and_format_df(p_raw.drop(columns=['Account', '_pid']), OMISSIONS)
        p_clean.columns = [f"Property: {c}" for c in p_clean.columns]
        p_clean["Account"] = p_raw["Account"]
        p_clean["_pid"] = p_raw["_pid"]

        i_raw = pd.read_sql("SELECT * FROM [TDS Insurance]", conn)
        i_clean = clean_and_format_df(i_raw, OMISSIONS)
        i_clean.columns = [f"Insurance: {c}" for c in i_clean.columns]
        i_clean["_pref"] = i_raw["PropRecID"]

        df4 = (
            pd.merge(p_clean, i_clean, left_on="_pid", right_on="_pref", how="left")
            .drop(columns=["_pid", "_pref"])
        )

        workbooks = {
            "1-Loans": df1,
            "2-Co-Borrowers": df2,
            "3-Fundings": df3,
            "4-Properties_&_Insurance": df4
        }

        for name, df in workbooks.items():
            path = os.path.join(output_folder, f"{name}.xlsx")
            df.to_excel(path, index=False)
            wb = load_workbook(path)
            ws = wb.active

            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = max(
                    len(str(c.value)) if c.value else 0 for c in col
                ) + 3

            url = MAPPING_LINKS.get(name)
            if url:
                defs = get_field_definitions(url)
                for cell in ws[1]:
                    key = str(cell.value).replace("Property: ", "").replace("Insurance: ", "")
                    if key in defs:
                        cell.comment = Comment(defs[key], "TMO Reference")

            wb.save(path)

        conn.close()
        messagebox.showinfo("Success", "Production export completed.")

    except Exception as e:
        messagebox.showerror("Error", str(e))


# -------------------------------
# Modern UI
# -------------------------------
def start_app():
    root = tk.Tk()
    root.title("TDS Excel Production Suite")
    root.geometry("520x360")
    root.configure(bg="#1e1e1e")

    style = ttk.Style(root)
    style.theme_use("clam")

    style.configure(".", background="#1e1e1e", foreground="#e6e6e6", font=("SF Pro", 12))
    style.configure("TEntry", fieldbackground="#2b2b2b")
    style.configure("TButton", padding=8)
    style.configure("Primary.TButton", background="#4a90e2", foreground="white")
    style.map("Primary.TButton", background=[("active", "#357abd")])

    frame = ttk.Frame(root, padding=20)
    frame.pack(fill="both", expand=True)

    ttk.Label(frame, text="Server").pack(anchor="w")
    server = ttk.Entry(frame, width=55)
    server.insert(0, "10.1.18.7")
    server.pack(pady=4)

    ttk.Label(frame, text="Database").pack(anchor="w", pady=(10, 0))
    database = ttk.Entry(frame, width=55)
    database.insert(0, "001-7782-000-Ramiro Blank Oct")
    database.pack(pady=4)

    ttk.Label(frame, text="Output Folder").pack(anchor="w", pady=(10, 0))
    folder_frame = ttk.Frame(frame)
    folder_frame.pack(fill="x")

    folder = ttk.Entry(folder_frame)
    folder.insert(0, os.getcwd())
    folder.pack(side="left", fill="x", expand=True)

    ttk.Button(
        folder_frame,
        text="Browse",
        command=lambda: folder.delete(0, tk.END) or folder.insert(0, filedialog.askdirectory())
    ).pack(side="left", padx=6)

    ttk.Button(
        frame,
        text="RUN PRODUCTION EXPORT",
        style="Primary.TButton",
        command=lambda: run_production_export(server.get(), database.get(), folder.get())
    ).pack(pady=30)

    root.mainloop()


if __name__ == "__main__":
    start_app()