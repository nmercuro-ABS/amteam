import pandas as pd
import pyodbc
import os
import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl.comments import Comment
from openpyxl import load_workbook
import warnings

# Suppress console warnings for a clean experience
warnings.filterwarnings('ignore', category=UserWarning)

# --- 1. FIELD MAPPING LINKS ---
MAPPING_LINKS = {
    '1-Loans': "https://help.themortgageoffice.com/knowledge/loan-field-mappings",
    '2-Co-Borrowers': "https://help.themortgageoffice.com/knowledge/co-borrower-field-mappings",
    '3-Fundings': "https://help.themortgageoffice.com/knowledge/what-are-the-vendor-field-mappings",
    '4-Properties_&_Insurance': "https://help.themortgageoffice.com/knowledge/what-are-the-properties-insurance-field-mappings"
}


def get_field_definitions(url):
    """Scrapes TMO help tables and preserves line breaks for Excel Notes."""
    defs = {}
    if not url: return defs
    try:
        response = requests.get(url, timeout=10)
        soup = BeautifulSoup(response.text, 'html.parser')
        for row in soup.find_all('tr'):
            cells = row.find_all(['td', 'th'])
            if len(cells) >= 2:
                field = cells[0].get_text(strip=True)
                desc_cell = cells[1]
                for tag in desc_cell.find_all(["br", "p", "li"]):
                    tag.insert_before("\n")
                raw_text = desc_cell.get_text()
                lines = [line.strip() for line in raw_text.splitlines() if line.strip()]
                defs[field] = "\n".join(lines)
    except:
        pass
    return defs


def clean_and_format_df(df, omissions):
    """Drops system fields, removes ALL empty/zero columns, then formats dates."""
    # 1. Drop RecID and Omissions
    recid_cols = [c for c in df.columns if 'recid' in c.lower()]
    df = df.drop(columns=recid_cols)
    df = df.drop(columns=[c for c in df.columns if c in omissions], errors='ignore')

    # 2. Strict Cleaning: Drop columns that are 100% empty, NULL, or 0
    invalid_vals = [0, 0.0, '0', '0.0', '', ' ', 'NULL', 'null', 'None', None]
    df = df.loc[:, ~(df.isin(invalid_vals)).all(axis=0)]

    # 3. Date Formatting: Only on columns that survived the cleaning
    for col in df.columns:
        # Check if column header suggests a date or if the data type is already datetime
        if 'date' in col.lower() or pd.api.types.is_datetime64_any_dtype(df[col]):
            try:
                # Convert to datetime and format to m/d/yyyy (no leading zeros)
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%#m/%#d/%Y')
                # Replace 'NaT' (failed conversions) with empty strings
                df[col] = df[col].replace('NaT', '')
            except:
                pass

    return df


def run_production_export(server, database, output_folder):
    try:
        OMISSIONS = [
            'SysCreatedDate', 'SysTimeStamp', 'SysRecStatus', 'SysCreatedBy',
            'BorrowerF', 'CompanyF', 'SortName', 'ByLastName', 'FundControl',
            'ACH_IndividualId', 'XML', 'LanguagePreference', 'InsuranceDocument'
        ]

        conn_str = (f"Driver={{ODBC Driver 17 for SQL Server}};Server={server};"
                    f"Database={database};Authentication=ActiveDirectoryInteractive;"
                    "Encrypt=yes;TrustServerCertificate=yes;")
        conn = pyodbc.connect(conn_str)

        if not os.path.exists(output_folder): os.makedirs(output_folder)

        # --- DATA EXTRACTION ---
        # 1-Loans
        df1 = pd.read_sql("SELECT * FROM [TDS Loans]", conn)
        df1 = clean_and_format_df(df1, OMISSIONS)
        # Manually append the two requested columns at the very end
        df1["ReserveBalance"], df1["ImpoundBalance"] = "", ""

        # 2-Co-Borrowers
        df2 = pd.read_sql(
            "SELECT l.Account, t.* FROM [TDS CoBorrowers] t LEFT JOIN [TDS Loans] l ON t.LoanRecID = l.RecID", conn)
        df2 = clean_and_format_df(df2, OMISSIONS)

        # 3-Fundings
        f_query = """SELECT l.Account AS Loan_Account, lend.Account AS Lender_Account, f.*, d.* FROM [TDS Funding] f 
                     INNER JOIN [TDS Draws] d ON f.RecID = d.FundingRecID 
                     LEFT JOIN [TDS Loans] l ON f.LoanRecID = l.RecID
                     LEFT JOIN [TDS Lenders] lend ON f.LenderRecID = lend.RecID"""
        df3 = pd.read_sql(f_query, conn)
        df3 = clean_and_format_df(df3, OMISSIONS)
        new_names = {c: c[:-3] + "_Pct" if c.endswith("Pct") else c[:-4] + "_Flat" if c.endswith("Flat") else c[
                                                                                                              :-3] + "_Min" if c.endswith(
            "Min") else c for c in df3.columns}
        df3.rename(columns=new_names, inplace=True)
        cols3 = ['Loan_Account', 'Lender_Account'] + [c for c in df3.columns if
                                                      c not in ['Loan_Account', 'Lender_Account']]
        df3 = df3[cols3]

        # 4-Properties & Insurance
        p_query = "SELECT l.Account, p.RecID as _pid, p.* FROM [TDS Properties] p LEFT JOIN [TDS Loans] l ON p.LoanRecID = l.RecID"
        df_p_raw = pd.read_sql(p_query, conn)
        df_p_clean = clean_and_format_df(df_p_raw.drop(columns=['Account', '_pid']), OMISSIONS)
        df_p_clean.columns = [f"Property: {c}" for c in df_p_clean.columns]
        df_p_clean['Account'], df_p_clean['_pid'] = df_p_raw['Account'], df_p_raw['_pid']
        df_i_raw = pd.read_sql("SELECT * FROM [TDS Insurance]", conn)
        df_i_clean = clean_and_format_df(df_i_raw, OMISSIONS)
        df_i_clean.columns = [f"Insurance: {c}" for c in df_i_clean.columns]
        df_i_clean['_pref'] = df_i_raw['PropRecID']
        df4 = pd.merge(df_p_clean, df_i_clean, left_on='_pid', right_on='_pref', how='left').drop(
            columns=['_pid', '_pref'])
        df4 = df4[['Account'] + [c for c in df4.columns if c != 'Account']]

        # 5-Escrow & 6-History
        df5 = pd.read_sql(
            "SELECT l.Account, t.* FROM [TDS Vouchers] t LEFT JOIN [TDS Loans] l ON t.LoanRecID = l.RecID", conn)
        df5 = clean_and_format_df(df5, OMISSIONS).rename(columns={'Account': 'Loan Account'})
        df6 = pd.read_sql(
            "SELECT l.Account, t.* FROM [TDS Loan History] t LEFT JOIN [TDS Loans] l ON t.LoanRecID = l.RecID", conn)
        df6 = clean_and_format_df(df6, OMISSIONS)

        workbooks = {'1-Loans': df1, '2-Co-Borrowers': df2, '3-Fundings': df3, '4-Properties_&_Insurance': df4,
                     '5-Escrow_Vouchers': df5, '6-Loan_History': df6}

        # --- EXPORT, AUTOFIT & NOTES ---
        for name, df in workbooks.items():
            path = os.path.join(output_folder, f"{name}.xlsx")
            df.to_excel(path, index=False)
            wb = load_workbook(path);
            ws = wb.active

            # AutoFit
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column[0].column_letter].width = (max_length + 2) * 1.15

            # Apply Notes
            url = MAPPING_LINKS.get(name)
            if url:
                defs = get_field_definitions(url)
                for cell in ws[1]:
                    lookup = str(cell.value).replace("Property: ", "").replace("Insurance: ", "")
                    if lookup in defs:
                        comm = Comment(defs[lookup], "TMO Reference")
                        comm.width, comm.height = 400, 200
                        cell.comment = comm
            wb.save(path)

        conn.close()
        messagebox.showinfo("Success", "Exports complete.")
    except Exception as e:
        messagebox.showerror("Error", str(e))


def start_app():
    root = tk.Tk();
    root.title("TDS Excel Production Suite");
    root.geometry("500x320")
    tk.Label(root, text="Server:").pack(pady=(10, 0))
    s_in = tk.Entry(root, width=55);
    s_in.insert(0, "10.1.18.7");
    s_in.pack()
    tk.Label(root, text="Database:").pack(pady=(5, 0))
    d_in = tk.Entry(root, width=55);
    d_in.insert(0, "001-7782-000-Ramiro Blank Oct");
    d_in.pack()
    tk.Label(root, text="Output Folder:").pack(pady=(5, 0))
    f_frame = tk.Frame(root);
    f_frame.pack()
    f_in = tk.Entry(f_frame, width=45);
    f_in.insert(0, os.getcwd());
    f_in.pack(side=tk.LEFT)
    tk.Button(f_frame, text="Browse",
              command=lambda: [f_in.delete(0, tk.END), f_in.insert(0, filedialog.askdirectory())]).pack(side=tk.LEFT)
    tk.Button(root, text="RUN PRODUCTION EXPORT", bg="#28a745", fg="white", font=("Arial", 11, "bold"), padx=20,
              pady=10, command=lambda: run_production_export(s_in.get(), d_in.get(), f_in.get())).pack(pady=25)
    root.mainloop()


if __name__ == "__main__": start_app()